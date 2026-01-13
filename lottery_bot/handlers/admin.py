"""Handlers dedicated to administrator actions."""
from __future__ import annotations

import asyncio
import os
import sys
import shutil
import tempfile
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.error import TelegramError
from telegram.ext import (
    CallbackContext,
    CallbackQueryHandler,
    CommandHandler,
    MessageHandler,
    filters,
)

from lottery_bot.keyboards import admin_menu_keyboard
from lottery_bot.storage import StorageManager


def register_admin_handlers(application) -> None:
    """Register admin command, message, and callback handlers."""
    settings = application.bot_data["settings"]
    admin_filter = filters.User(user_id=[settings.admin_id])

    admin_cancel_handler = CommandHandler("cancel", admin_cancel, filters=admin_filter)
    application.add_handler(admin_cancel_handler, group=-1)
    application.add_handler(CommandHandler("admin", admin_home, filters=admin_filter))
    
    # Main menu handlers
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^🏠 Bosh sahifa$"), admin_home_dashboard))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^📊 Statistika$"), admin_stats))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^⏳ Kutilayotgan to'lovlar$"), admin_pending_payments))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^✅ Tasdiqlangan to'lovlar$"), admin_list_approved))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^✉️ Xabar yuborish$"), admin_broadcast_entry))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^👥 Foydalanuvchilar$"), admin_users_list))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^📡 Kanal boshqaruvi$"), admin_subscription_entry))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^📥 Excel eksport$"), admin_export_excel))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("^⚙️ Bot sozlamlari$"), admin_settings_entry))
    
    # Legacy handlers (for backward compatibility)
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Xabar yuborish"), admin_broadcast_entry))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Statistika"), admin_stats))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Tasdiqlanganlarni bekor qilish"), admin_list_approved))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Bot sozlamlari"), admin_settings_entry))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Kanal boshqaruvi"), admin_subscription_entry))
    application.add_handler(MessageHandler(admin_filter & filters.Regex("Excel eksport"), admin_export_excel))
    
    # Callback handlers
    application.add_handler(CallbackQueryHandler(admin_decision, pattern=r"^(approve|reject):"))
    application.add_handler(CallbackQueryHandler(admin_subscription_toggle, pattern="^subscription:toggle$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_refresh, pattern="^subscription:refresh$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_close, pattern="^subscription:close$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_list, pattern="^subscription:list$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_invite_link, pattern="^subscription:invite_link$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_preview, pattern="^subscription:preview$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_add, pattern="^subscription:add$"))
    application.add_handler(
        CallbackQueryHandler(admin_subscription_prompt_remove, pattern="^subscription:prompt_remove$")
    )
    application.add_handler(CallbackQueryHandler(admin_subscription_remove, pattern=r"^subscription:remove:"))
    application.add_handler(CallbackQueryHandler(admin_subscription_cancel_input, pattern="^subscription:cancel_input$"))
    application.add_handler(CallbackQueryHandler(admin_subscription_no_channels, pattern="^subscription:no_channels$"))
    application.add_handler(
        CallbackQueryHandler(admin_subscription_edit_message, pattern="^subscription:edit_message$")
    )
    application.add_handler(CallbackQueryHandler(admin_cancel_approved, pattern=r"^approved:cancel:"))
    application.add_handler(CallbackQueryHandler(admin_approved_close, pattern="^approved:close$"))
    application.add_handler(CallbackQueryHandler(admin_settings_restart, pattern="^settings:restart$"))
    application.add_handler(CallbackQueryHandler(admin_settings_backup, pattern="^settings:backup$"))
    application.add_handler(CallbackQueryHandler(admin_settings_restore, pattern="^settings:restore$"))
    application.add_handler(CallbackQueryHandler(admin_settings_clear_data, pattern="^settings:clear_data$"))
    application.add_handler(CallbackQueryHandler(admin_settings_clear_confirm, pattern="^settings:clear_confirm$"))
    application.add_handler(CallbackQueryHandler(admin_settings_change_card, pattern="^settings:change_card$"))
    application.add_handler(CallbackQueryHandler(admin_settings_change_manager, pattern="^settings:change_manager$"))
    application.add_handler(CallbackQueryHandler(admin_settings_cancel_input, pattern="^settings:cancel_input$"))
    application.add_handler(CallbackQueryHandler(admin_start_message_entry_cb, pattern="^settings:start_edit$"))
    application.add_handler(CallbackQueryHandler(admin_start_message_cancel, pattern="^cancel_start_message$"))
    application.add_handler(CallbackQueryHandler(admin_game_info_message_entry_cb, pattern="^settings:game_info_edit$"))
    application.add_handler(CallbackQueryHandler(admin_game_info_message_cancel, pattern="^cancel_game_info_message$"))
    application.add_handler(CallbackQueryHandler(admin_game_info_message_reset, pattern="^reset_game_info_message$"))
    application.add_handler(CallbackQueryHandler(admin_broadcast_cancel, pattern="^cancel_broadcast$"))
    application.add_handler(CallbackQueryHandler(admin_settings_close, pattern="^settings:close$"))
    application.add_handler(CallbackQueryHandler(admin_pending_page, pattern=r"^pending:page:"))
    application.add_handler(CallbackQueryHandler(admin_pending_close, pattern="^pending:close$"))
    application.add_handler(CallbackQueryHandler(admin_users_page, pattern=r"^users:page:"))
    application.add_handler(CallbackQueryHandler(admin_users_close, pattern="^users:close$"))
    
    subscription_input_handler = MessageHandler(admin_filter & ~filters.COMMAND, admin_subscription_text_input)
    subscription_input_handler.block = False
    application.add_handler(subscription_input_handler, group=5)

    settings_input_handler = MessageHandler(admin_filter & ~filters.COMMAND, admin_settings_text_input)
    settings_input_handler.block = False
    application.add_handler(settings_input_handler, group=6)

    mode_router_handler = MessageHandler(admin_filter & ~filters.COMMAND, admin_active_mode_router)
    mode_router_handler.block = False
    application.add_handler(mode_router_handler, group=7)


def _format_money(value: int | float) -> str:
    return f"{value:,.0f}".replace(",", " ")


def _format_decimal(value: float) -> str:
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _format_money_decimal(value: float) -> str:
    formatted = f"{value:,.2f}".replace(",", " ")
    return formatted.rstrip("0").rstrip(".")


def _settings_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("🔄 Botni qayta ishga tushirish", callback_data="settings:restart")],
            [
                InlineKeyboardButton("💾 Zaxira olish", callback_data="settings:backup"),
                InlineKeyboardButton("📥 Zaxirani tiklash", callback_data="settings:restore"),
            ],
            [InlineKeyboardButton("💳 Karta raqamini almashtirish", callback_data="settings:change_card")],
            [InlineKeyboardButton("👤 Menejer kontaktini almashtirish", callback_data="settings:change_manager")],
            [InlineKeyboardButton("✏️ Start xabarini tahrirlash", callback_data="settings:start_edit")],
            [InlineKeyboardButton("ℹ️ O'yin haqida xabarini tahrirlash", callback_data="settings:game_info_edit")],
            [InlineKeyboardButton("🗑 Bazani tozalash", callback_data="settings:clear_data")],
            [InlineKeyboardButton("❌ Yopish", callback_data="settings:close")],
        ]
    )


def _subscription_management_keyboard(enabled: bool, has_channels: bool, channels_count: int = 0) -> InlineKeyboardMarkup:
    """Build subscription management keyboard."""
    toggle_icon = "🟢" if enabled else "🔴"
    status = "Yoqilgan" if enabled else "O'chirilgan"
    toggle_text = f"{toggle_icon} Obuna: {status}"
    
    buttons = [
        [InlineKeyboardButton(toggle_text, callback_data="subscription:toggle")],
        [
            InlineKeyboardButton("➕ Kanal qo'shish", callback_data="subscription:add"),
            InlineKeyboardButton(f"🗑 O'chirish ({channels_count})", callback_data="subscription:prompt_remove" if has_channels else "subscription:no_channels"),
        ],
        [
            InlineKeyboardButton("📋 Ro'yxat", callback_data="subscription:list"),
            InlineKeyboardButton("🔗 Taklif havolasi", callback_data="subscription:invite_link"),
        ],
        [
            InlineKeyboardButton("✏️ Xabarni tahrirlash", callback_data="subscription:edit_message"),
            InlineKeyboardButton("👁 Ko'rish", callback_data="subscription:preview"),
        ],
        [
            InlineKeyboardButton("🔄 Yangilash", callback_data="subscription:refresh"),
            InlineKeyboardButton("❌ Yopish", callback_data="subscription:close"),
        ],
    ]
    return InlineKeyboardMarkup(buttons)


def _format_channel_list(channels, detailed: bool = False) -> str:
    """Format channel list for display."""
    if not channels:
        return "📭 Hali kanal qo'shilmagan."
    
    lines = []
    for idx, channel in enumerate(channels, start=1):
        title = channel.get("title") or channel.get("id") or "Kanal"
        link = channel.get("link")
        channel_id = channel.get("id", "")
        
        if detailed:
            lines.append(f"<b>{idx}. {title}</b>")
            if link:
                lines.append(f"   🔗 {link}")
            if channel_id:
                lines.append(f"   🆔 {channel_id}")
            lines.append("")
        else:
            if link:
                lines.append(f"{idx}. {title}")
            else:
                lines.append(f"{idx}. {title} (havolasiz)")
    
    return "\n".join(lines)


def _build_subscription_summary(config, notice: Optional[str] = None) -> tuple[str, InlineKeyboardMarkup]:
    """Build subscription management panel."""
    enabled = config.get("enabled", False)
    channels = config.get("channels", [])
    
    status_icon = "🟢" if enabled else "🔴"
    status_text = "Yoqilgan" if enabled else "O'chirilgan"
    
    lines = [
        "📡 <b>Kanal boshqaruvi</b>",
        "━━━━━━━━━━━━━━━━━━━━",
        "",
        f"📊 <b>Holat:</b>",
        f"   {status_icon} Majburiy obuna: <b>{status_text}</b>",
        f"   📺 Kanallar soni: <b>{len(channels)}</b>",
        "",
        "━━━━━━━━━━━━━━━━━━━━",
        "📋 <b>Kanallar:</b>",
        _format_channel_list(channels),
        "━━━━━━━━━━━━━━━━━━━━",
    ]
    
    if notice:
        lines.extend(["", f"💬 {notice}"])
    
    keyboard = _subscription_management_keyboard(enabled, bool(channels), len(channels))
    return "\n".join(lines), keyboard


def _set_subscription_message_ref(context: CallbackContext, chat_id: int, message_id: int) -> None:
    context.user_data["subscription_message_ref"] = (chat_id, message_id)


def _build_approved_summary(approved) -> tuple[str, Optional[InlineKeyboardMarkup]]:
    if not approved:
        return (
            "✅ Hozircha bekor qilish uchun tasdiqlangan to'lovlar yo'q.",
            None,
        )

    lines = ["✅ Tasdiqlangan to'lovlar (bekor qilish mumkin):"]
    buttons = []
    for item in approved:
        pid = item.get("purchase_id")
        user_display = item.get("full_name") or f"ID {item.get('user_id')}"
        tickets = item.get("tickets", [])
        amount = _format_money(item.get("amount", 0))
        lines.append(
            f"• {pid} | {user_display} | 🎟 {len(tickets)} ta | 💰 {amount} so'm"
        )
        buttons.append(
            [InlineKeyboardButton(f"↩️ Bekor qilish — {pid}", callback_data=f"approved:cancel:{pid}")]
        )

    buttons.append([InlineKeyboardButton("❌ Yopish", callback_data="approved:close")])
    return "\n".join(lines), InlineKeyboardMarkup(buttons)


async def _refresh_subscription_message_from_ref(
    context: CallbackContext,
    *,
    notice: Optional[str] = None,
) -> None:
    ref = context.user_data.get("subscription_message_ref")
    if not ref:
        return
    chat_id, message_id = ref
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    summary, keyboard = _build_subscription_summary(config, notice=notice)
    try:
        await context.bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=summary,
            reply_markup=keyboard,
        )
    except TelegramError:
        pass


# ==================== ADMIN HOME & DASHBOARD ====================

async def admin_home(update: Update, context: CallbackContext) -> None:
    """Show the admin navigation menu."""
    await update.message.reply_text(
        "👋 Admin paneliga xush kelibsiz!\n\n"
        "Quyidagi menyudan kerakli bo'limni tanlang:",
        reply_markup=admin_menu_keyboard()
    )


async def admin_home_dashboard(update: Update, context: CallbackContext) -> None:
    """Show admin dashboard with quick stats."""
    storage: StorageManager = context.application.bot_data["storage"]
    stats = await storage.get_detailed_stats()
    
    progress = int((stats['tickets_sold'] / stats['total_tickets']) * 100) if stats['total_tickets'] > 0 else 0
    progress_bar = "▓" * (progress // 10) + "░" * (10 - progress // 10)
    
    dashboard = f"""
🏠 <b>Admin Dashboard</b>

━━━━━━━━━━━━━━━━━━━━
📊 <b>Tezkor statistika</b>
━━━━━━━━━━━━━━━━━━━━

🎟 <b>Chiptalar:</b>
{progress_bar} {progress}%
Sotilgan: {stats['tickets_sold']} / {stats['total_tickets']}

💰 <b>Daromad:</b> {_format_money(stats['total_revenue'])} so'm

👥 <b>Foydalanuvchilar:</b> {stats['total_users']} ta
📥 <b>Kutilayotgan:</b> {stats['pending_count']} ta

━━━━━━━━━━━━━━━━━━━━
"""
    
    await update.message.reply_text(
        dashboard,
        parse_mode="HTML",
        reply_markup=admin_menu_keyboard()
    )


# ==================== PENDING PAYMENTS ====================

async def admin_pending_payments(update: Update, context: CallbackContext) -> None:
    """Show pending payments list with pagination."""
    storage: StorageManager = context.application.bot_data["storage"]
    pending = await storage.list_pending()
    
    if not pending:
        await update.message.reply_text(
            "⏳ Hozircha kutilayotgan to'lovlar yo'q.",
            reply_markup=admin_menu_keyboard()
        )
        return
    
    page = 0
    text, keyboard = _build_pending_list(pending, page)
    await update.message.reply_text(text, reply_markup=keyboard, parse_mode="HTML")


def _build_pending_list(pending: list, page: int = 0, per_page: int = 5) -> tuple[str, InlineKeyboardMarkup]:
    """Build pending payments list with pagination."""
    total = len(pending)
    total_pages = (total + per_page - 1) // per_page
    start = page * per_page
    end = start + per_page
    items = pending[start:end]
    
    lines = [
        f"⏳ <b>Kutilayotgan to'lovlar</b>",
        f"📊 Jami: {total} ta | Sahifa: {page + 1}/{max(1, total_pages)}",
        "━━━━━━━━━━━━━━━━━━━━",
        ""
    ]
    
    for item in items:
        user_display = item.get("full_name") or f"ID: {item.get('user_id')}"
        username = f"@{item['username']}" if item.get('username') else "—"
        phone = item.get('phone_number') or "—"
        amount = _format_money(item.get("amount", 0))
        quantity = item.get("quantity", 1)
        created = item.get("created_at", "")[:10] if item.get("created_at") else "—"
        
        lines.append(f"👤 <b>{user_display}</b>")
        lines.append(f"   📱 {username} | 📞 {phone}")
        lines.append(f"   🎟 {quantity} ta | 💰 {amount} so'm")
        lines.append(f"   📅 {created}")
        lines.append("")
    
    buttons = []
    nav_buttons = []
    
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("◀️ Oldingi", callback_data=f"pending:page:{page - 1}"))
    if page < total_pages - 1:
        nav_buttons.append(InlineKeyboardButton("Keyingi ▶️", callback_data=f"pending:page:{page + 1}"))
    
    if nav_buttons:
        buttons.append(nav_buttons)
    buttons.append([InlineKeyboardButton("❌ Yopish", callback_data="pending:close")])
    
    return "\n".join(lines), InlineKeyboardMarkup(buttons)


async def admin_pending_page(update: Update, context: CallbackContext) -> None:
    """Handle pending payments pagination."""
    query = update.callback_query
    await query.answer()
    
    parts = query.data.split(":")
    page = int(parts[2]) if len(parts) > 2 else 0
    
    storage: StorageManager = context.application.bot_data["storage"]
    pending = await storage.list_pending()
    
    if not pending:
        await query.edit_message_text("⏳ Kutilayotgan to'lovlar yo'q.")
        return
    
    text, keyboard = _build_pending_list(pending, page)
    try:
        await query.edit_message_text(text, reply_markup=keyboard, parse_mode="HTML")
    except TelegramError:
        pass


async def admin_pending_close(update: Update, context: CallbackContext) -> None:
    """Close pending payments list."""
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_text("⏳ Kutilayotgan to'lovlar ro'yxati yopildi.")
    except TelegramError:
        pass


# ==================== USERS LIST ====================

async def admin_users_list(update: Update, context: CallbackContext) -> None:
    """Show users list with pagination."""
    storage: StorageManager = context.application.bot_data["storage"]
    users = await storage.list_all_users()
    
    if not users:
        await update.message.reply_text(
            "👥 Hozircha foydalanuvchilar yo'q.",
            reply_markup=admin_menu_keyboard()
        )
        return
    
    page = 0
    text, keyboard = _build_users_list(users, page)
    await update.message.reply_text(text, reply_markup=keyboard, parse_mode="HTML")


def _build_users_list(users: list, page: int = 0, per_page: int = 10) -> tuple[str, InlineKeyboardMarkup]:
    """Build users list with pagination."""
    total = len(users)
    total_pages = (total + per_page - 1) // per_page
    start = page * per_page
    end = start + per_page
    items = users[start:end]
    
    lines = [
        f"👥 <b>Foydalanuvchilar ro'yxati</b>",
        f"📊 Jami: {total} ta | Sahifa: {page + 1}/{max(1, total_pages)}",
        "━━━━━━━━━━━━━━━━━━━━",
        ""
    ]
    
    for idx, user in enumerate(items, start=start + 1):
        name = user.get("full_name") or "Noma'lum"
        username = f"@{user['username']}" if user.get('username') else "—"
        phone = user.get('phone_number') or "—"
        tickets = user.get('total_tickets', 0)
        spent = _format_money(user.get('total_spent', 0))
        
        lines.append(f"{idx}. <b>{name}</b>")
        lines.append(f"   📱 {username} | 📞 {phone}")
        lines.append(f"   🎟 {tickets} ta | 💰 {spent} so'm")
        lines.append("")
    
    buttons = []
    nav_buttons = []
    
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("◀️ Oldingi", callback_data=f"users:page:{page - 1}"))
    if page < total_pages - 1:
        nav_buttons.append(InlineKeyboardButton("Keyingi ▶️", callback_data=f"users:page:{page + 1}"))
    
    if nav_buttons:
        buttons.append(nav_buttons)
    buttons.append([InlineKeyboardButton("❌ Yopish", callback_data="users:close")])
    
    return "\n".join(lines), InlineKeyboardMarkup(buttons)


async def admin_users_page(update: Update, context: CallbackContext) -> None:
    """Handle users list pagination."""
    query = update.callback_query
    await query.answer()
    
    parts = query.data.split(":")
    page = int(parts[2]) if len(parts) > 2 else 0
    
    storage: StorageManager = context.application.bot_data["storage"]
    users = await storage.list_all_users()
    
    if not users:
        await query.edit_message_text("👥 Foydalanuvchilar yo'q.")
        return
    
    text, keyboard = _build_users_list(users, page)
    try:
        await query.edit_message_text(text, reply_markup=keyboard, parse_mode="HTML")
    except TelegramError:
        pass


async def admin_users_close(update: Update, context: CallbackContext) -> None:
    """Close users list."""
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_text("👥 Foydalanuvchilar ro'yxati yopildi.")
    except TelegramError:
        pass


# ==================== STATISTICS ====================


async def admin_stats(update: Update, context: CallbackContext) -> None:
    """Deliver detailed analytics for the admin."""
    storage: StorageManager = context.application.bot_data["storage"]
    stats = await storage.get_detailed_stats()
    
    # Progress bar
    progress = int((stats['tickets_sold'] / stats['total_tickets']) * 100) if stats['total_tickets'] > 0 else 0
    progress_bar = "▓" * (progress // 10) + "░" * (10 - progress // 10)

    summary_lines = [
        "📊 <b>Batafsil statistika</b>",
        "━━━━━━━━━━━━━━━━━━━━",
        "",
        "👥 <b>Foydalanuvchilar:</b>",
        f"   • Jami: <b>{stats['total_users']}</b>",
        f"   • 24 soat ichida faol: {stats['active_users_24h']}",
        f"   • 24 soat ichida yangi: {stats['new_users_24h']}",
        "",
        "🎟 <b>Chiptalar:</b>",
        f"   {progress_bar} {progress}%",
        f"   • Sotilgan: <b>{stats['tickets_sold']}</b> / {stats['total_tickets']}",
        f"   • Qolgan: <b>{stats['remaining_tickets']}</b>",
        f"   • O'rtacha chipta/foydalanuvchi: {_format_decimal(stats['avg_tickets_per_user'])}",
        "",
        "💰 <b>Moliya:</b>",
        f"   • Jami daromad: <b>{_format_money(stats['total_revenue'])} so'm</b>",
        f"   • O'rtacha to'lov: {_format_money_decimal(stats['avg_spend_per_user'])} so'm",
        "",
        "📋 <b>To'lovlar:</b>",
        f"   • Jami: {stats['total_purchases']}",
        f"   • ✅ Tasdiqlangan: {stats['approved_count']}",
        f"   • ❌ Rad etilgan: {stats['rejected_count']}",
        f"   • ⏳ Kutilayotgan: {stats['pending_count']} (≈{_format_money(stats['pending_amount'])} so'm)",
    ]

    if stats["top_users"]:
        summary_lines.append("")
        summary_lines.append("🏆 <b>TOP 5 qatnashchilar:</b>")
        medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
        for idx, entry in enumerate(stats["top_users"]):
            display_name = entry.get("full_name") or f"ID {entry['user_id']}"
            if entry.get("username"):
                display_name += f" (@{entry['username']})"
            medal = medals[idx] if idx < len(medals) else f"{idx + 1}."
            summary_lines.append(
                f"   {medal} {display_name}"
            )
            summary_lines.append(
                f"       🎟 {entry['tickets']} ta | 💰 {_format_money(entry['spent'])} so'm"
            )
    
    summary_lines.append("")
    summary_lines.append("━━━━━━━━━━━━━━━━━━━━")

    await update.message.reply_text(
        "\n".join(summary_lines), 
        reply_markup=admin_menu_keyboard(),
        parse_mode="HTML"
    )


# ==================== APPROVED PAYMENTS ====================

async def admin_list_approved(update: Update, context: CallbackContext) -> None:
    """List approved purchases with an option to cancel them."""
    storage: StorageManager = context.application.bot_data["storage"]
    approved = await storage.list_approved()
    text, markup = _build_approved_summary(approved)
    await update.message.reply_text(text, reply_markup=markup or admin_menu_keyboard())


async def admin_subscription_entry(update: Update, context: CallbackContext) -> None:
    """Open the subscription management popup."""
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    summary, keyboard = _build_subscription_summary(config)
    sent = await update.message.reply_text(summary, reply_markup=keyboard)
    _set_subscription_message_ref(context, sent.chat_id, sent.message_id)


async def admin_settings_entry(update: Update, context: CallbackContext) -> None:
    """Show bot settings actions."""
    storage: StorageManager = context.application.bot_data["storage"]
    card_number = await storage.get_card_number()
    manager_contact = await storage.get_manager_contact()
    text = (
        "⚙️ Bot sozlamlari\n"
        f"• Joriy karta: {card_number or 'kiritilmagan'}\n"
        f"• Menejer: {manager_contact or 'belgilangan emas'}\n"
        "• Backup va restartni shu yerda boshqarishingiz mumkin."
    )
    await update.message.reply_text(text, reply_markup=_settings_keyboard())


async def _edit_subscription_menu(query, context: CallbackContext, notice: Optional[str] = None) -> None:
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    summary, keyboard = _build_subscription_summary(config, notice=notice)
    try:
        await query.edit_message_text(summary, reply_markup=keyboard)
    except TelegramError:
        pass
    else:
        if query.message:
            _set_subscription_message_ref(context, query.message.chat_id, query.message.message_id)


async def admin_subscription_refresh(update: Update, context: CallbackContext) -> None:
    """Refresh the subscription overview from inline button."""
    query = update.callback_query
    await query.answer()
    await _edit_subscription_menu(query, context)


async def admin_settings_close(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_text("⚙️ Sozlamalar yopildi.")
    except TelegramError:
        pass


async def admin_settings_restart(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer("Bot qayta ishga tushirilmoqda...", show_alert=True)
    await query.edit_message_text("🔄 Bot qayta ishga tushirilmoqda...")
    # Restart current process.
    os.execl(sys.executable, sys.executable, *sys.argv)


async def admin_settings_backup(update: Update, context: CallbackContext) -> None:
    """Create and send database backup."""
    query = update.callback_query
    await query.answer("💾 Zaxira nusxa tayyorlanmoqda...")
    
    storage: StorageManager = context.application.bot_data["storage"]
    src_path = getattr(storage, "_path", None)
    if not src_path or not os.path.exists(src_path):
        await query.answer("Backup uchun fayl topilmadi.", show_alert=True)
        return

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"lottery_backup_{timestamp}.json"
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
    try:
        shutil.copy(src_path, temp_file.name)
        temp_file.close()
        
        # Get stats for caption
        stats = await storage.get_detailed_stats()
        
        caption = (
            f"💾 <b>Zaxira nusxa</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📅 Sana: {timestamp.replace('_', ' ')}\n"
            f"👥 Foydalanuvchilar: {stats['total_users']}\n"
            f"🎟 Sotilgan chiptalar: {stats['tickets_sold']}\n"
            f"💰 Daromad: {_format_money(stats['total_revenue'])} so'm\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📥 Tiklash uchun bu faylni botga yuboring."
        )
        
        with open(temp_file.name, "rb") as handle:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=handle,
                filename=filename,
                caption=caption,
                parse_mode="HTML",
            )
        
        await query.answer("✅ Zaxira nusxa yuborildi!", show_alert=True)
    except Exception as e:
        await query.answer(f"❌ Xatolik: {str(e)[:100]}", show_alert=True)
    finally:
        try:
            os.remove(temp_file.name)
        except OSError:
            pass


async def admin_settings_restore(update: Update, context: CallbackContext) -> None:
    """Prompt admin to send backup file for restoration."""
    query = update.callback_query
    await query.answer()
    
    context.user_data["settings_mode"] = "restore_backup"
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "📥 <b>Zaxira nusxani tiklash</b>\n\n"
            "⚠️ <b>Diqqat!</b> Bu amal joriy barcha ma'lumotlarni o'chirib, "
            "zaxira nusxadagi ma'lumotlar bilan almashtiradi.\n\n"
            "Davom etish uchun zaxira faylini (.json) yuboring.\n"
            "Bekor qilish uchun tugmadan foydalaning."
        ),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("❌ Bekor qilish", callback_data="settings:cancel_input")]]
        ),
    )


async def admin_settings_clear_data(update: Update, context: CallbackContext) -> None:
    """Confirm before clearing all data."""
    query = update.callback_query
    await query.answer()
    
    storage: StorageManager = context.application.bot_data["storage"]
    stats = await storage.get_detailed_stats()
    
    await query.edit_message_text(
        text=(
            "🗑 <b>Bazani tozalash</b>\n\n"
            "⚠️ <b>OGOHLANTIRISH!</b>\n"
            "Bu amal quyidagi ma'lumotlarni O'CHIRIB TASHLAYDI:\n\n"
            f"👥 Foydalanuvchilar: {stats['total_users']} ta\n"
            f"🎟 Sotilgan chiptalar: {stats['tickets_sold']} ta\n"
            f"💰 Daromad: {_format_money(stats['total_revenue'])} so'm\n"
            f"📋 To'lovlar: {stats['total_purchases']} ta\n\n"
            "❗ Bu amalni ortga qaytarib bo'lmaydi!\n"
            "Davom etishdan oldin zaxira nusxa oling."
        ),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🗑 HA, TOZALASH", callback_data="settings:clear_confirm")],
            [InlineKeyboardButton("❌ Bekor qilish", callback_data="settings:close")],
        ]),
    )


async def admin_settings_clear_confirm(update: Update, context: CallbackContext) -> None:
    """Actually clear all data after confirmation."""
    query = update.callback_query
    await query.answer("🗑 Baza tozalanmoqda...", show_alert=True)
    
    storage: StorageManager = context.application.bot_data["storage"]
    
    try:
        await storage.reset_all_data()
        await query.edit_message_text(
            "✅ <b>Baza muvaffaqiyatli tozalandi!</b>\n\n"
            "Barcha ma'lumotlar o'chirildi. Bot yangi holatda.",
            parse_mode="HTML",
        )
        
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="⚙️ Bot sozlamalari",
            reply_markup=_settings_keyboard(),
        )
    except Exception as e:
        await query.edit_message_text(f"❌ Xatolik yuz berdi: {str(e)[:200]}")


async def admin_settings_handle_restore(update: Update, context: CallbackContext) -> None:
    """Handle backup file upload for restoration."""
    if context.user_data.get("settings_mode") != "restore_backup":
        return
    
    message = update.message
    if not message.document:
        await message.reply_text("❗ Iltimos, .json fayl yuboring.")
        return
    
    if not message.document.file_name.endswith(".json"):
        await message.reply_text("❗ Faqat .json formatdagi fayl qabul qilinadi.")
        return
    
    try:
        # Download the file
        file = await context.bot.get_file(message.document.file_id)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
        await file.download_to_drive(temp_file.name)
        temp_file.close()
        
        # Validate JSON
        import json
        with open(temp_file.name, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Check required keys
        required_keys = ["available_tickets", "users", "approved", "pending"]
        missing_keys = [k for k in required_keys if k not in data]
        if missing_keys:
            await message.reply_text(
                f"❗ Noto'g'ri format. Quyidagi kalitlar topilmadi: {', '.join(missing_keys)}"
            )
            os.remove(temp_file.name)
            return
        
        # Restore data
        storage: StorageManager = context.application.bot_data["storage"]
        await storage.restore_from_backup(temp_file.name)
        
        context.user_data.pop("settings_mode", None)
        
        # Get new stats
        stats = await storage.get_detailed_stats()
        
        await message.reply_text(
            f"✅ <b>Zaxira nusxa muvaffaqiyatli tiklandi!</b>\n\n"
            f"📊 <b>Tiklangan ma'lumotlar:</b>\n"
            f"👥 Foydalanuvchilar: {stats['total_users']}\n"
            f"🎟 Sotilgan chiptalar: {stats['tickets_sold']}\n"
            f"💰 Daromad: {_format_money(stats['total_revenue'])} so'm\n\n"
            f"⚠️ Botni qayta ishga tushirish tavsiya etiladi.",
            parse_mode="HTML",
            reply_markup=admin_menu_keyboard(),
        )
        
        os.remove(temp_file.name)
        
    except json.JSONDecodeError:
        await message.reply_text("❗ Fayl noto'g'ri JSON formatda.")
    except Exception as e:
        await message.reply_text(f"❌ Xatolik yuz berdi: {str(e)[:200]}")


async def admin_settings_change_card(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["settings_mode"] = "card_number"
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "💳 Yangi karta raqamini yuboring.\n"
            "Misol: 9860 1234 5678 9012\n"
            "Bekor qilish uchun tugmadan foydalaning."
        ),
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("Bekor qilish", callback_data="settings:cancel_input")]]
        ),
    )


async def admin_settings_change_manager(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["settings_mode"] = "manager_contact"
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "👤 Yangi menejer username ni yuboring.\n"
        ),
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("Bekor qilish", callback_data="settings:cancel_input")]]
        ),
    )


async def admin_subscription_toggle(update: Update, context: CallbackContext) -> None:
    """Toggle mandatory subscription state."""
    query = update.callback_query
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    new_state = not config.get("enabled", False)
    await storage.set_subscription_enabled(new_state)
    await _edit_subscription_menu(query, context, notice="✅ Holat yangilandi.")
    status_text = "Majburiy obuna yoqildi." if new_state else "Majburiy obuna o'chirildi."
    await query.answer(status_text, show_alert=True)


async def admin_subscription_close(update: Update, context: CallbackContext) -> None:
    """Close the subscription popup."""
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_text("📡 Kanal boshqaruvi yopildi.")
    except TelegramError:
        pass


async def admin_subscription_add(update: Update, context: CallbackContext) -> None:
    """Prepare to add a new subscription channel."""
    query = update.callback_query
    await query.answer()
    context.user_data["subscription_mode"] = "add"
    if query.message:
        _set_subscription_message_ref(context, query.message.chat_id, query.message.message_id)
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "➕ Kanal qo'shish uchun kanal username yoki havolasini yuboring, yoki kanaldan xabarni "
            "forward qiling. Bekor qilish uchun tugmadan foydalaning."
        ),
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("Bekor qilish", callback_data="subscription:cancel_input")]]
        ),
    )


async def admin_subscription_prompt_remove(update: Update, context: CallbackContext) -> None:
    """Show removable channels."""
    query = update.callback_query
    await query.answer()
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    channels = config.get("channels", [])
    if not channels:
        await query.answer("Kanal qo'shilmagan.", show_alert=True)
        return

    buttons = [
        [
            InlineKeyboardButton(
                f"❌ {channel.get('title') or channel.get('id')}", callback_data=f"subscription:remove:{channel.get('id')}"
            )
        ]
        for channel in channels
    ]
    buttons.append([InlineKeyboardButton("⬅️ Ortga", callback_data="subscription:refresh")])
    try:
        await query.edit_message_text(
            "❌ O'chirish uchun kanalni tanlang:", reply_markup=InlineKeyboardMarkup(buttons)
        )
    except TelegramError:
        pass


async def admin_subscription_cancel_input(update: Update, context: CallbackContext) -> None:
    """Cancel the current subscription input (add/edit) and return to menu."""
    query = update.callback_query
    await query.answer("Bekor qilindi", show_alert=True)
    context.user_data.pop("subscription_mode", None)
    await _edit_subscription_menu(query, context, notice="❌ Jarayon bekor qilindi.")


async def admin_subscription_remove(update: Update, context: CallbackContext) -> None:
    """Remove a selected subscription channel."""
    query = update.callback_query
    parts = query.data.split(":", maxsplit=2)
    channel_id = parts[2] if len(parts) > 2 else ""
    storage: StorageManager = context.application.bot_data["storage"]
    removed = await storage.remove_subscription_channel(channel_id)
    notice = "✅ Kanal o'chirildi." if removed else "ℹ️ Kanal topilmadi."
    await _edit_subscription_menu(query, context, notice=notice)
    await query.answer(notice, show_alert=True)


async def admin_subscription_edit_message(update: Update, context: CallbackContext) -> None:
    """Prompt admin to edit subscription reminder text."""
    query = update.callback_query
    await query.answer()
    context.user_data["subscription_mode"] = "edit_message"
    if query.message:
        _set_subscription_message_ref(context, query.message.chat_id, query.message.message_id)
    current = await context.application.bot_data["storage"].get_subscription_message()
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "📝 Yangi majburiy obuna xabarini yuboring. Matnda {channels} o'zgaruvchisi bo'lishi kerak.\n"
            "Bekor qilish uchun /cancel yuboring.\n"
            "\nJoriy xabar:\n" + current
        ),
    )


async def _resolve_channel(update: Update, context: CallbackContext) -> Optional[dict]:
    """Extract channel information from the admin's message."""
    message = update.message
    chat = message.forward_from_chat
    if chat is None:
        text = (message.text or "").strip()
        if not text:
            await message.reply_text("❗ Kanal username yoki havolasini yuboring.")
            return None
        identifier = text.split()[0]
        if "t.me/" in identifier:
            identifier = identifier.split("t.me/")[-1]
        identifier = identifier.strip()
        if identifier.startswith("https://") or identifier.startswith("http://"):
            identifier = identifier.split("/")[-1]
        if not identifier.startswith("@") and not identifier.startswith("-100"):
            identifier = f"@{identifier}" if not identifier.startswith("+") else identifier
        try:
            chat = await context.bot.get_chat(identifier)
        except TelegramError:
            await message.reply_text("❗ Kanal topilmadi. Username yoki havolasini tekshiring.")
            return None

    if chat.type not in ("channel", "supergroup"):
        await message.reply_text("❗ Faqat kanal yoki supergruppani yuboring.")
        return None

    channel_id = str(chat.id)
    title = chat.title or getattr(chat, "full_name", None) or chat.username or channel_id
    username = getattr(chat, "username", None)
    link = None
    if username:
        link = f"https://t.me/{username}"
    elif getattr(chat, "invite_link", None):
        link = chat.invite_link

    return {"channel_id": channel_id, "title": title, "link": link}


async def admin_subscription_text_input(update: Update, context: CallbackContext) -> None:
    """Handle follow-up text for subscription management actions."""
    mode = context.user_data.get("subscription_mode")
    if not mode:
        return

    storage: StorageManager = context.application.bot_data["storage"]

    if mode == "add":
        channel = await _resolve_channel(update, context)
        if not channel:
            return
        await storage.add_subscription_channel(channel["channel_id"], channel["title"], channel.get("link"))
        context.user_data.pop("subscription_mode", None)
        await update.message.reply_text(
            "✅ Kanal qo'shildi. Iltimos, botni ushbu kanalga administrator sifatida qo'shganingizga ishonch hosil qiling.",
        )
        await _refresh_subscription_message_from_ref(context, notice="✅ Kanal qo'shildi.")
    elif mode == "edit_message":
        text = (update.message.text or "").strip()
        if not text:
            await update.message.reply_text("❗ Matn bo'sh bo'lishi mumkin emas. Qayta yuboring.")
            return
        try:
            await storage.set_subscription_message(text)
        except ValueError as exc:
            await update.message.reply_text(f"❗ {exc}")
            return
        context.user_data.pop("subscription_mode", None)
        await update.message.reply_text("✅ Majburiy obuna xabari yangilandi.")
        await _refresh_subscription_message_from_ref(context, notice="✅ Xabar yangilandi.")


async def admin_settings_text_input(update: Update, context: CallbackContext) -> None:
    """Handle text inputs for settings flows (e.g., card number)."""
    mode = context.user_data.get("settings_mode")
    if not mode:
        return

    if mode == "card_number":
        card = (update.message.text or "").strip()
        if not card:
            await update.message.reply_text("❗ Karta raqami bo'sh bo'lmasligi kerak.")
            return
        storage: StorageManager = context.application.bot_data["storage"]
        await storage.set_card_number(card)
        context.user_data.pop("settings_mode", None)
        await update.message.reply_text(
            f"✅ Karta raqami yangilandi: {card}", reply_markup=admin_menu_keyboard()
        )
        await context.bot.send_message(
            chat_id=update.message.chat_id,
            text="⚙️ Bot sozlamalari yangilandi.",
            reply_markup=_settings_keyboard(),
        )
    elif mode == "manager_contact":
        contact = (update.message.text or "").strip()
        if not contact:
            await update.message.reply_text("❗ Username bo'sh bo'lmasligi kerak.")
            return
        if not contact.startswith("@"):  # normalize
            contact = "@" + contact
        storage: StorageManager = context.application.bot_data["storage"]
        await storage.set_manager_contact(contact)
        context.user_data.pop("settings_mode", None)
        await update.message.reply_text(
            f"✅ Menejer kontakti yangilandi: {contact}", reply_markup=admin_menu_keyboard()
        )
        await context.bot.send_message(
            chat_id=update.message.chat_id,
            text="⚙️ Bot sozlamalari yangilandi.",
            reply_markup=_settings_keyboard(),
        )


async def admin_settings_cancel_input(update: Update, context: CallbackContext) -> None:
    """Cancel current settings input (card/manager)."""
    query = update.callback_query
    await query.answer("Bekor qilindi", show_alert=True)
    context.user_data.pop("settings_mode", None)
    try:
        await query.edit_message_text("⚙️ Amal bekor qilindi.")
    except TelegramError:
        pass


async def admin_subscription_list(update: Update, context: CallbackContext) -> None:
    """Send the current channel list in chat."""
    query = update.callback_query
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    channels = config.get("channels", [])
    if not channels:
        await query.answer("Hali kanal qo'shilmagan.", show_alert=True)
        return

    text_lines = ["📋 Kanallar ro'yxati:"]
    for idx, channel in enumerate(channels, start=1):
        title = channel.get("title") or channel.get("id") or "Kanal"
        link = channel.get("link")
        if link:
            text_lines.append(f"{idx}. {title} — {link}")
        else:
            text_lines.append(f"{idx}. {title}")

    await query.answer("Kanallar ro'yxati yuborildi.")
    await context.bot.send_message(chat_id=query.message.chat_id, text="\n".join(text_lines))


async def admin_subscription_invite_link(update: Update, context: CallbackContext) -> None:
    """Show invite links for all channels."""
    query = update.callback_query
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    channels = config.get("channels", [])
    
    if not channels:
        await query.answer("Hali kanal qo'shilmagan.", show_alert=True)
        return
    
    text_lines = ["🔗 <b>Kanal taklif havolalari:</b>", ""]
    for idx, channel in enumerate(channels, start=1):
        title = channel.get("title") or channel.get("id") or "Kanal"
        link = channel.get("link")
        if link:
            text_lines.append(f"{idx}. <b>{title}</b>\n   {link}")
        else:
            text_lines.append(f"{idx}. <b>{title}</b>\n   ⚠️ Havola mavjud emas")
        text_lines.append("")
    
    await query.answer("Taklif havolalari yuborildi.")
    await context.bot.send_message(
        chat_id=query.message.chat_id, 
        text="\n".join(text_lines),
        parse_mode="HTML"
    )


async def admin_subscription_preview(update: Update, context: CallbackContext) -> None:
    """Show a preview of the subscription check message as users see it."""
    query = update.callback_query
    storage: StorageManager = context.application.bot_data["storage"]
    config = await storage.get_subscription_config()
    channels = config.get("channels", [])
    custom_message = config.get("message", "")
    
    if not channels:
        await query.answer("Hali kanal qo'shilmagan.", show_alert=True)
        return
    
    # Build preview message
    preview_lines = [
        "👁 <b>Foydalanuvchi ko'radigan xabar:</b>",
        "━━━━━━━━━━━━━━━━━━━━",
        ""
    ]
    
    if custom_message:
        preview_lines.append(custom_message)
    else:
        preview_lines.append("⚠️ Botdan foydalanish uchun quyidagi kanallarga obuna bo'ling:")
    
    preview_lines.append("")
    for idx, channel in enumerate(channels, start=1):
        title = channel.get("title") or channel.get("id") or "Kanal"
        link = channel.get("link")
        if link:
            preview_lines.append(f"📢 <a href='{link}'>{title}</a>")
        else:
            preview_lines.append(f"📢 {title}")
    
    preview_lines.extend([
        "",
        "━━━━━━━━━━━━━━━━━━━━",
        "<i>✅ Obuna bo'lgandan so'ng tugmani bosing.</i>"
    ])
    
    await query.answer("Ko'rinish yuborildi.")
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="\n".join(preview_lines),
        parse_mode="HTML",
        disable_web_page_preview=True
    )


async def admin_subscription_no_channels(update: Update, context: CallbackContext) -> None:
    """Inform admin that there are no channels to remove."""
    query = update.callback_query
    await query.answer("Hali kanal qo'shilmagan.", show_alert=True)


async def admin_approved_close(update: Update, context: CallbackContext) -> None:
    """Close the approved-list message."""
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_text("✅ Tasdiqlangan to'lovlar ro'yxati yopildi.")
    except TelegramError:
        pass


async def admin_cancel_approved(update: Update, context: CallbackContext) -> None:
    """Cancel an approved purchase and refund tickets."""
    query = update.callback_query
    parts = query.data.split(":", maxsplit=2)
    purchase_id = parts[2] if len(parts) > 2 else ""
    storage: StorageManager = context.application.bot_data["storage"]
    purchase = await storage.cancel_approved_purchase(purchase_id)
    if not purchase:
        await query.answer("Topilmadi yoki allaqachon bekor qilingan.", show_alert=True)
        return

    tickets = purchase.get("tickets", [])
    amount = _format_money(purchase.get("amount", 0))
    ticket_list = ", ".join(str(t) for t in sorted(tickets)) if tickets else "-"

    await query.answer("Bekor qilindi", show_alert=True)

    refreshed = await storage.list_approved()
    text, markup = _build_approved_summary(refreshed)
    try:
        await query.edit_message_text(text, reply_markup=markup)
    except TelegramError:
        pass

    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "↩️ Tasdiqlangan chek bekor qilindi.\n"
            f"Chek: {purchase_id}\n"
            f"Foydalanuvchi: {purchase.get('full_name') or purchase.get('user_id')}\n"
            f"🎟 Chiptalar: {ticket_list}\n"
            f"💰 To'lov: {amount} so'm"
        ),
        reply_markup=admin_menu_keyboard(),
    )

    # Notify user about cancellation if possible.
    user_id = purchase.get("user_id")
    if user_id:
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    "❗️ Sizning tasdiqlangan chiptangiz bekor qilindi.\n"
                    f"Chek ID: {purchase_id}\n"
                    f"🎟 Chiptalar: {ticket_list}\n"
                    "Savol bo'lsa, administrator bilan bog'laning."
                ),
            )
        except TelegramError:
            pass


async def admin_export_excel(update: Update, context: CallbackContext) -> None:
    """Generate and send an Excel report of approved purchases."""
    storage: StorageManager = context.application.bot_data["storage"]
    rows = await storage.get_ticket_export_rows()
    if not rows:
        await update.message.reply_text(
            "📭 Hozircha eksport qilish uchun tasdiqlangan to'lovlar yo'q.",
            reply_markup=admin_menu_keyboard(),
        )
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Chiptalar"
    headers = [
        "Purchase ID",
        "Foydalanuvchi",
        "Username",
        "Telefon",
        "Chipta soni",
        "Chipta raqamlari",
        "To'lov (so'm)",
        "Tasdiqlangan vaqt",
    ]
    sheet.append(headers)

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 22
    for col_idx, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align

    # Wider, clearer columns
    widths = [18, 22, 18, 16, 14, 28, 16, 26]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in rows:
        tickets = ", ".join(str(ticket) for ticket in sorted(row.get("tickets", [])))
        sheet.append(
            [
                row.get("purchase_id"),
                row.get("full_name"),
                ("@" + row["username"]) if row.get("username") else "",
                row.get("phone_number") or "",
                row.get("quantity", 0),
                tickets,
                row.get("amount", 0),
                row.get("resolved_at") or "",
            ]
        )

    # Align text for readability
    center_cols = {1, 5, 7, 8}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        workbook.save(temp_file.name)
        temp_file.close()
        with open(temp_file.name, "rb") as handle:
            await update.message.reply_document(
                document=handle,
                filename="lottery_export.xlsx",
                caption="📥 Tasdiqlangan chiptalar bo'yicha hisobot tayyor.",
            )
    finally:
        try:
            os.remove(temp_file.name)
        except OSError:
            pass

async def admin_decision(update: Update, context: CallbackContext) -> None:
    """Handle approval or rejection callbacks."""
    query = update.callback_query
    await query.answer()

    settings = context.application.bot_data["settings"]
    if query.from_user.id != settings.admin_id:
        await query.answer(text="Bu tugma faqat admin uchun.", show_alert=True)
        return

    action, purchase_id = query.data.split(":", maxsplit=1)
    storage: StorageManager = context.application.bot_data["storage"]

    if not await storage.is_pending(purchase_id):
        await _edit_admin_message(query, "ℹ️ Bu chek allaqachon ko'rib chiqilgan.")
        return

    if action == "approve":
        tickets, purchase = await storage.approve_purchase(purchase_id)
        if not tickets:
            await query.answer(text="Yetarli chipta qolmadi yoki purchase topilmadi.", show_alert=True)
            return

        ticket_list = ", ".join(str(t) for t in sorted(tickets))
        amount = _format_money(purchase.get("amount", 0))

        await _edit_admin_message(
            query,
            f"✅ Tasdiqlandi\n🎟 Chiptalar: {ticket_list}\n💰 To'lov: {amount} so'm",
        )

        await context.bot.send_message(
            chat_id=purchase["user_id"],
            text=(
                "🎉 Tabriklaymiz! To'lovingiz muvaffaqiyatli tasdiqlandi.\n"
                f"🎟 Sizga biriktirilgan chiptalar: {ticket_list}\n"
                "🙏 Ishtirokingiz uchun rahmat, omad yor bo'lsin!"
            ),
        )

        await context.bot.send_message(
            chat_id=settings.admin_id,
            text=f"✅ Tasdiqlandi: {purchase['full_name']} — {ticket_list} (💰 {amount} so'm)",
        )
    else:
        purchase = await storage.reject_purchase(purchase_id)
        if not purchase:
            await _edit_admin_message(query, "ℹ️ Bu chek allaqachon ko'rib chiqilgan.")
            return

        await _edit_admin_message(query, "❌ Chek rad etildi.")
        manager_contact = await storage.get_manager_contact()
        contact_username = manager_contact.lstrip("@") or "menejer_1w"
        contact_url = f"https://t.me/{contact_username}"
        await context.bot.send_message(
            chat_id=purchase["user_id"],
            text=(
                "❌ Kechirasiz, to'lov tasdiqlanmadi.\n"
                "Iltimos, ma'lumotlarni tekshirib, qayta yuboring."
            ),
            reply_markup=InlineKeyboardMarkup(
                [
                    [InlineKeyboardButton("admin bilan bog'lanish", url=contact_url)]
                ]
            ),
        )


async def admin_broadcast_entry(update: Update, context: CallbackContext) -> None:
    """Prompt admin for broadcast content."""
    if context.user_data.get("broadcast_mode") == "awaiting_content":
        await update.message.reply_text("ℹ️ Hozirda yuboriladigan xabarni kutyapman. Iltimos, xabarni yuboring yoki bekor qiling.")
        return

    context.user_data["broadcast_mode"] = "awaiting_content"
    if update.message:
        context.user_data["broadcast_ignore_message_id"] = update.message.message_id
    await update.message.reply_text(
        (
            "✉️ Yuboriladigan xabarni yuboring. Matn, rasm yoki video (caption bilan) qo'llab-quvvatlanadi.\n"
            "Bekor qilish uchun tugmadan foydalaning."
        ),
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("Bekor qilish", callback_data="cancel_broadcast")]]
        ),
    )


async def admin_broadcast_handle_content(update: Update, context: CallbackContext) -> None:
    """Send broadcast message to all known users when mode is active."""
    if context.user_data.get("broadcast_mode") != "awaiting_content":
        return

    ignore_id = context.user_data.pop("broadcast_ignore_message_id", None)
    if ignore_id and update.message and update.message.message_id == ignore_id:
        return

    message = update.message
    if message.photo:
        payload = {
            "type": "photo",
            "file_id": message.photo[-1].file_id,
            "caption": (message.caption or "").strip(),
        }
    elif message.video:
        payload = {
            "type": "video",
            "file_id": message.video.file_id,
            "caption": (message.caption or "").strip(),
        }
    else:
        text = (message.text or "").strip()
        if not text:
            await update.message.reply_text("❗ Xabar bo'sh bo'lishi mumkin emas. Qaytadan yuboring.")
            return
        payload = {"type": "text", "text": text}

    storage: StorageManager = context.application.bot_data["storage"]
    user_ids = await storage.list_user_ids()
    if not user_ids:
        context.user_data.pop("broadcast_mode", None)
        await update.message.reply_text(
            "📭 Hozircha xabar yuboriladigan foydalanuvchi mavjud emas.",
            reply_markup=admin_menu_keyboard(),
        )
        return

    await update.message.reply_text(
        f"✉️ Xabar yuborilmoqda... (jami {len(user_ids)} foydalanuvchi)", reply_markup=admin_menu_keyboard()
    )

    delivered = 0
    failed = 0
    for user_id in user_ids:
        try:
            if payload["type"] == "text":
                await context.bot.send_message(chat_id=user_id, text=payload["text"])
            elif payload["type"] == "photo":
                await context.bot.send_photo(
                    chat_id=user_id,
                    photo=payload["file_id"],
                    caption=payload.get("caption") or None,
                )
            elif payload["type"] == "video":
                await context.bot.send_video(
                    chat_id=user_id,
                    video=payload["file_id"],
                    caption=payload.get("caption") or None,
                )
            delivered += 1
        except TelegramError:
            failed += 1
            continue
        await asyncio.sleep(0.05)

    context.user_data.pop("broadcast_mode", None)
    await update.message.reply_text(
        f"✅ Yuborildi: {delivered} ta\n⚠️ Yuborilmadi: {failed} ta",
        reply_markup=admin_menu_keyboard(),
    )


async def admin_broadcast_cancel(update: Update, context: CallbackContext) -> None:
    """Cancel the broadcast flow from inline button."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("broadcast_mode", None)
    context.user_data.pop("broadcast_ignore_message_id", None)
    await query.edit_message_text("✉️ Xabar yuborish bekor qilindi.")


async def admin_start_message_entry_cb(update: Update, context: CallbackContext) -> None:
    """Start-message edit flow triggered from settings inline button."""
    query = update.callback_query
    await query.answer()
    context.user_data["start_edit_mode"] = True
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "✏️ Start xabarining yangi matnini yoki media (caption bilan) yuboring.\n"
            "Bekor qilish uchun tugmadan foydalaning."
        ),
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("Bekor qilish", callback_data="cancel_start_message")]]
        ),
    )


async def admin_start_message_handle_input(update: Update, context: CallbackContext) -> None:
    """Persist the new start message template when editing mode is active."""
    if not context.user_data.get("start_edit_mode"):
        return

    message = update.message
    if message.photo:
        text = (message.caption or "").strip()
        media = {"type": "photo", "file_id": message.photo[-1].file_id}
    elif message.video:
        text = (message.caption or "").strip()
        media = {"type": "video", "file_id": message.video.file_id}
    else:
        text = (message.text or "").strip()
        media = None

    if not text:
        await update.message.reply_text("❗ Matn bo'sh bo'lishi mumkin emas. Qayta kiriting.")
        return

    storage: StorageManager = context.application.bot_data["storage"]
    settings = context.application.bot_data["settings"]

    try:
        await storage.set_start_message(text=text, media=media)
    except ValueError as exc:
        await update.message.reply_text(f"❗ {exc}")
        return

    remaining = await storage.remaining_tickets()
    preview = await storage.render_start_content(
        prize=settings.prize_name,
        total_tickets=settings.total_tickets,
        remaining_tickets=remaining,
        ticket_price=_format_money(settings.ticket_price),
    )

    if preview.get("media") and preview["media"].get("type") == "photo":
        await update.message.reply_photo(
            photo=preview["media"]["file_id"],
            caption="✅ Start xabari yangilandi:\n\n" + preview["text"],
            reply_markup=admin_menu_keyboard(),
        )
    elif preview.get("media") and preview["media"].get("type") == "video":
        await update.message.reply_video(
            video=preview["media"]["file_id"],
            caption="✅ Start xabari yangilandi:\n\n" + preview["text"],
            reply_markup=admin_menu_keyboard(),
        )
    else:
        await update.message.reply_text(
            "✅ Start xabari yangilandi. Joriy ko'rinish:\n\n" + preview["text"],
            reply_markup=admin_menu_keyboard(),
        )
    context.user_data.pop("start_edit_mode", None)


async def admin_start_message_cancel(update: Update, context: CallbackContext) -> None:
    """Cancel the start message editing flow via inline button."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("start_edit_mode", None)
    await query.edit_message_text("✏️ Start xabarini tahrirlash bekor qilindi.")


async def admin_game_info_message_entry_cb(update: Update, context: CallbackContext) -> None:
    """Begin game-info message editing from settings."""
    query = update.callback_query
    await query.answer()
    context.user_data["game_info_edit_mode"] = True
    storage: StorageManager = context.application.bot_data["storage"]
    current = await storage.get_game_info_message()
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=(
            "ℹ️ Yangi 'O'yin haqida' xabarini yuboring. Matnda quyidagi o'zgaruvchilardan foydalanishingiz mumkin:\n"
            "{prize}, {total_tickets}, {sold_tickets}, {remaining_tickets}, {ticket_price}.\n"
            "Standart holatga qaytarish yoki bekor qilish uchun pastdagi tugmalardan foydalaning.\n\n"
            "Joriy xabar:\n"
            f"{current}"
        ),
        reply_markup=InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("♻️ Standart holatga qaytarish", callback_data="reset_game_info_message")],
                [InlineKeyboardButton("Bekor qilish", callback_data="cancel_game_info_message")],
            ]
        ),
    )


async def admin_game_info_message_handle_input(update: Update, context: CallbackContext) -> None:
    """Persist the updated game-info message template."""
    if not context.user_data.get("game_info_edit_mode"):
        return

    text = (update.message.text or "").strip()
    if not text:
        await update.message.reply_text("❗ Faqat matn yuboring. Bo'sh xabar qabul qilinmaydi.")
        return

    storage: StorageManager = context.application.bot_data["storage"]
    settings = context.application.bot_data["settings"]

    try:
        await storage.set_game_info_message(text)
    except ValueError as exc:
        await update.message.reply_text(f"❗ {exc}")
        return

    preview = await storage.render_game_info_message(
        prize=settings.prize_name,
        total_tickets=settings.total_tickets,
        ticket_price=_format_money(settings.ticket_price),
    )

    await update.message.reply_text(
        "✅ 'O'yin haqida' xabari yangilandi. Joriy ko'rinish:\n\n" + preview,
        reply_markup=admin_menu_keyboard(),
    )
    context.user_data.pop("game_info_edit_mode", None)


async def admin_game_info_message_cancel(update: Update, context: CallbackContext) -> None:
    """Cancel game-info editing flow via inline button."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("game_info_edit_mode", None)
    await query.edit_message_text("ℹ️ O'yin haqida xabarini tahrirlash bekor qilindi.")


async def admin_game_info_message_reset(update: Update, context: CallbackContext) -> None:
    """Restore the game-info message to its default template."""
    query = update.callback_query
    await query.answer()
    storage: StorageManager = context.application.bot_data["storage"]
    settings = context.application.bot_data["settings"]
    await storage.reset_game_info_message()
    preview = await storage.render_game_info_message(
        prize=settings.prize_name,
        total_tickets=settings.total_tickets,
        ticket_price=_format_money(settings.ticket_price),
    )

    context.user_data.pop("game_info_edit_mode", None)

    try:
        await query.edit_message_text("♻️ 'O'yin haqida' xabari standart holatga qaytarildi.")
    except TelegramError:
        pass

    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="✅ Joriy ko'rinish:\n\n" + preview,
        reply_markup=admin_menu_keyboard(),
    )


async def admin_cancel(update: Update, context: CallbackContext) -> None:
    """Fallback handler to exit admin flows via /cancel."""
    context.user_data.pop("subscription_mode", None)
    context.user_data.pop("settings_mode", None)
    context.user_data.pop("broadcast_mode", None)
    context.user_data.pop("start_edit_mode", None)
    context.user_data.pop("game_info_edit_mode", None)
    await update.message.reply_text("❌ Jarayon bekor qilindi.", reply_markup=admin_menu_keyboard())


async def admin_active_mode_router(update: Update, context: CallbackContext) -> None:
    """Route incoming admin messages to active modes (broadcast/start edit/restore)."""
    if context.user_data.get("broadcast_mode") == "awaiting_content":
        await admin_broadcast_handle_content(update, context)
        return
    if context.user_data.get("start_edit_mode"):
        await admin_start_message_handle_input(update, context)
        return
    if context.user_data.get("game_info_edit_mode"):
        await admin_game_info_message_handle_input(update, context)
        return
    if context.user_data.get("settings_mode") == "restore_backup":
        await admin_settings_handle_restore(update, context)
        return


async def _edit_admin_message(query, text: str) -> None:
    """Edit the admin's message caption or text."""
    message = query.message
    if message.photo or message.document:
        await query.edit_message_caption(caption=text, reply_markup=None)
    else:
        await query.edit_message_text(text=text, reply_markup=None)

